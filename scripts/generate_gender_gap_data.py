from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


FIELDS = {
    "median": 4,
}

AGE_LABELS = {
    "16-17": "16–17",
    "18-21": "18–21",
    "22-29": "22–29",
    "30-39": "30–39",
    "40-49": "40–49",
    "50-59": "50–59",
    "60+": "60+",
}

AGE_BANDS = ["18-21", "22-29", "30-39", "40-49", "50-59", "60+"]

OCCUPATION_SHORT_LABELS = {
    "1": "Managers",
    "2": "Professional",
    "3": "Associate prof.",
    "4": "Admin & secretarial",
    "5": "Skilled trades",
    "6": "Care & leisure",
    "7": "Sales & service",
    "8": "Operatives",
    "9": "Elementary",
}

INDUSTRY_SHORT_LABELS = {
    "A": "Agriculture",
    "B": "Mining",
    "C": "Manufacturing",
    "D": "Utilities",
    "E": "Water & waste",
    "F": "Construction",
    "G": "Wholesale & retail",
    "H": "Transport",
    "I": "Hospitality",
    "J": "Information & comms",
    "K": "Finance",
    "L": "Property",
    "M": "Professional",
    "N": "Admin support",
    "O": "Public admin",
    "P": "Education",
    "Q": "Health & social care",
    "R": "Arts & recreation",
    "S": "Other services",
    "T": "Households",
    "U": "Extraterritorial",
}

REGION_SHORT_LABELS = {
    "K02000001": "UK",
    "E12000001": "North East",
    "E12000002": "North West",
    "E12000003": "Yorks & Humber",
    "E12000004": "East Midlands",
    "E12000005": "West Midlands",
    "E12000006": "East of England",
    "E12000007": "London",
    "E12000008": "South East",
    "E12000009": "South West",
    "W92000004": "Wales",
    "S92000003": "Scotland",
    "N92000002": "N. Ireland",
}

SECTOR_SHORT_LABELS = {
    "public-sector": "Public",
    "private-sector": "Private",
    "non-profit-body-or-mutual-association": "Non-profit",
    "not-classified": "Unclassified",
}

TABLE_FILES = {
    "age": "PROV - Age Group Table 6.6a   Hourly pay - Excluding overtime 2025.xlsx",
    "occupation": "PROV - Occupation SOC20 (2) Table 2.6a   Hourly pay - Excluding overtime 2025.xlsx",
    "industry": "PROV - SIC07 Industry (2) SIC2007 Table 4.6a   Hourly pay - Excluding overtime 2025.xlsx",
    "region": "PROV - Work Region Age Table WGOR Age.6a   Hourly pay - Excluding overtime 2025.xlsx",
    "sector": "PROV - PubPriv Table 13.6a   Hourly pay - Excluding overtime 2025.xlsx",
    "ageOccupation": "PROV - Age by Occupation SOC20 (2) Table 20.6a   Hourly pay - Excluding overtime 2025.xlsx",
}


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def compute_gap(male_median, female_median):
    if male_median in (None, 0) or female_median is None:
        return None
    return round(((male_median - female_median) / male_median) * 100, 1)


def make_gap_row(row_id, label, short_label, male_median, female_median):
    overall_median = None
    if male_median is not None and female_median is not None:
        overall_median = round((male_median + female_median) / 2, 2)
    return {
        "id": row_id,
        "label": label,
        "shortLabel": short_label,
        "maleMedian": male_median,
        "femaleMedian": female_median,
        "median": overall_median,
        "gapPct": compute_gap(male_median, female_median),
    }


def load_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    return workbook[sheet_name]


def build_age_rows(path: Path, male_sheet_name: str, female_sheet_name: str):
    male_sheet = load_sheet(path, male_sheet_name)
    female_sheet = load_sheet(path, female_sheet_name)
    rows = []
    overall = None

    male_map = {}
    female_map = {}

    for row in male_sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label == "All employees":
            overall = {
                "maleMedian": clean_value(row[FIELDS["median"] - 1]),
            }
            continue
        if label.startswith(("All employees", "Not Classified", "a  ", "b  ", "KEY", "The quality", "Source:")):
            continue
        normalized = AGE_LABELS.get(label.replace("b", "").strip(), label)
        male_map[normalized] = clean_value(row[FIELDS["median"] - 1])

    for row in female_sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        if not isinstance(label, str):
            continue
        label = label.strip()
        if label == "All employees":
            overall = {
                **(overall or {}),
                "femaleMedian": clean_value(row[FIELDS["median"] - 1]),
            }
            continue
        if label.startswith(("All employees", "Not Classified", "a  ", "b  ", "KEY", "The quality", "Source:")):
            continue
        normalized = AGE_LABELS.get(label.replace("b", "").strip(), label)
        female_map[normalized] = clean_value(row[FIELDS["median"] - 1])

    for label in ["16–17", "18–21", "22–29", "30–39", "40–49", "50–59", "60+"]:
        if label not in male_map and label not in female_map:
            continue
        rows.append(make_gap_row(label, label, label, male_map.get(label), female_map.get(label)))

    overall_row = make_gap_row(
        "all-employees",
        "All employees",
        "All employees",
        overall.get("maleMedian") if overall else None,
        overall.get("femaleMedian") if overall else None,
    )

    return rows, overall_row


def build_keyed_rows(path: Path, male_sheet_name: str, female_sheet_name: str, valid_id, short_labels, label_transform=None, id_transform=None):
    male_sheet = load_sheet(path, male_sheet_name)
    female_sheet = load_sheet(path, female_sheet_name)

    male_map = {}
    female_map = {}

    for row in male_sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        if row_id is None:
            continue
        row_id = str(row_id).strip()
        if not valid_id(row_id):
            continue
        mapped_id = id_transform(row_id) if id_transform else row_id
        label = clean_value(row[0])
        label = label_transform(label) if label_transform else label
        male_map[mapped_id] = {
            "label": label,
            "median": clean_value(row[FIELDS["median"] - 1]),
        }

    for row in female_sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        if row_id is None:
            continue
        row_id = str(row_id).strip()
        if not valid_id(row_id):
            continue
        mapped_id = id_transform(row_id) if id_transform else row_id
        label = clean_value(row[0])
        label = label_transform(label) if label_transform else label
        female_map[mapped_id] = {
            "label": label,
            "median": clean_value(row[FIELDS["median"] - 1]),
        }

    rows = []
    for row_id in male_map.keys():
        if row_id not in female_map:
            continue
        label = male_map[row_id]["label"]
        rows.append(
            make_gap_row(
                row_id,
                label,
                short_labels.get(row_id, label),
                male_map[row_id]["median"],
                female_map[row_id]["median"],
            )
        )

    return rows


def build_sector_rows(path: Path, male_sheet_name: str, female_sheet_name: str):
    male_sheet = load_sheet(path, male_sheet_name)
    female_sheet = load_sheet(path, female_sheet_name)

    def label_to_id(label: str):
        return re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")

    male_map = {}
    female_map = {}

    for row in male_sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        if not isinstance(label, str):
            continue
        row_id = label_to_id(label)
        if row_id not in SECTOR_SHORT_LABELS:
            continue
        male_map[row_id] = {
            "label": label.strip(),
            "median": clean_value(row[FIELDS["median"] - 1]),
        }

    for row in female_sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        if not isinstance(label, str):
            continue
        row_id = label_to_id(label)
        if row_id not in SECTOR_SHORT_LABELS:
            continue
        female_map[row_id] = {
            "label": label.strip(),
            "median": clean_value(row[FIELDS["median"] - 1]),
        }

    return [
        make_gap_row(
            row_id,
            male_map[row_id]["label"],
            SECTOR_SHORT_LABELS[row_id],
            male_map[row_id]["median"],
            female_map[row_id]["median"],
        )
        for row_id in male_map.keys()
        if row_id in female_map
    ]


def build_age_occupation_rows(path: Path, male_sheet_name: str, female_sheet_name: str):
    male_sheet = load_sheet(path, male_sheet_name)
    female_sheet = load_sheet(path, female_sheet_name)
    age_bands = {age_band: {} for age_band in AGE_BANDS}
    pattern = re.compile(r"^(18-21|22-29|30-39|40-49|50-59|60\+)\s+(.*)$")

    for row in male_sheet.iter_rows(min_row=13, values_only=True):
        row_id = clean_value(row[1])
        row_id = str(row_id).strip() if row_id is not None else ""
        if len(row_id) != 1 or not row_id.isdigit():
            continue
        label = " ".join(str(row[0]).split())
        match = pattern.match(label)
        if not match:
            continue
        age_band, occupation_label = match.groups()
        age_bands[age_band][row_id] = {
            "label": occupation_label,
            "maleMedian": clean_value(row[FIELDS["median"] - 1]),
        }

    for row in female_sheet.iter_rows(min_row=13, values_only=True):
        row_id = clean_value(row[1])
        row_id = str(row_id).strip() if row_id is not None else ""
        if len(row_id) != 1 or not row_id.isdigit():
            continue
        label = " ".join(str(row[0]).split())
        match = pattern.match(label)
        if not match:
            continue
        age_band, occupation_label = match.groups()
        entry = age_bands[age_band].setdefault(row_id, {"label": occupation_label})
        entry["femaleMedian"] = clean_value(row[FIELDS["median"] - 1])

    output = {}
    for age_band, rows in age_bands.items():
        output[age_band] = [
            make_gap_row(
                row_id,
                payload["label"],
                OCCUPATION_SHORT_LABELS.get(row_id, payload["label"]),
                payload.get("maleMedian"),
                payload.get("femaleMedian"),
            )
            for row_id, payload in rows.items()
        ]
    return output


def sort_rows(rows):
    return sorted(rows, key=lambda row: row["id"])


def main():
    if len(sys.argv) != 7:
        raise SystemExit("Usage: generate_gender_gap_data.py <table6_dir> <table2_dir> <table4_dir> <table15_dir> <table13_dir> <table20_dir>")

    table6_dir, table2_dir, table4_dir, table15_dir, table13_dir, table20_dir = [Path(arg) for arg in sys.argv[1:]]

    age_all_rows, age_all_overall = build_age_rows(table6_dir / TABLE_FILES["age"], "Male", "Female")
    age_ft_rows, age_ft_overall = build_age_rows(table6_dir / TABLE_FILES["age"], "Male Full-Time", "Female Full-Time")
    age_data = {
        "all": age_all_rows,
        "ft": age_ft_rows,
    }
    gap_benchmarks = {
        "all": age_all_overall,
        "ft": age_ft_overall,
    }

    occupation_data = {
        "all": sort_rows(build_keyed_rows(
            table2_dir / TABLE_FILES["occupation"],
            "Male",
            "Female",
            valid_id=lambda row_id: len(row_id) == 1 and row_id.isdigit(),
            short_labels=OCCUPATION_SHORT_LABELS,
        )),
        "ft": sort_rows(build_keyed_rows(
            table2_dir / TABLE_FILES["occupation"],
            "Male Full-Time",
            "Female Full-Time",
            valid_id=lambda row_id: len(row_id) == 1 and row_id.isdigit(),
            short_labels=OCCUPATION_SHORT_LABELS,
        )),
    }

    industry_data = {
        "all": sort_rows(build_keyed_rows(
            table4_dir / TABLE_FILES["industry"],
            "Male",
            "Female",
            valid_id=lambda row_id: len(row_id) == 1 and row_id.isalpha(),
            short_labels=INDUSTRY_SHORT_LABELS,
            label_transform=lambda label: str(label).strip().title(),
        )),
        "ft": sort_rows(build_keyed_rows(
            table4_dir / TABLE_FILES["industry"],
            "Male Full-Time",
            "Female Full-Time",
            valid_id=lambda row_id: len(row_id) == 1 and row_id.isalpha(),
            short_labels=INDUSTRY_SHORT_LABELS,
            label_transform=lambda label: str(label).strip().title(),
        )),
    }

    region_data = {
        "all": sort_rows(build_keyed_rows(
            table15_dir / TABLE_FILES["region"],
            "Male",
            "Female",
            valid_id=lambda row_id: row_id in REGION_SHORT_LABELS,
            short_labels=REGION_SHORT_LABELS,
        )),
        "ft": sort_rows(build_keyed_rows(
            table15_dir / TABLE_FILES["region"],
            "Male Full-Time",
            "Female Full-Time",
            valid_id=lambda row_id: row_id in REGION_SHORT_LABELS,
            short_labels=REGION_SHORT_LABELS,
        )),
    }

    sector_data = {
        "all": build_sector_rows(table13_dir / TABLE_FILES["sector"], "Male", "Female"),
        "ft": build_sector_rows(table13_dir / TABLE_FILES["sector"], "Male Full-Time", "Female Full-Time"),
    }

    age_occupation_data = {
        "all": build_age_occupation_rows(table20_dir / TABLE_FILES["ageOccupation"], "Male", "Female"),
        "ft": build_age_occupation_rows(table20_dir / TABLE_FILES["ageOccupation"], "Male Full-Time", "Female Full-Time"),
    }

    output = [
        "// Generated from ONS ASHE 2025 hourly excluding overtime workbooks.",
        f"export const GAP_BENCHMARKS = {json.dumps(gap_benchmarks, ensure_ascii=False, indent=2)};",
        "",
        f"export const AGE_GAP_DATA = {json.dumps(age_data, ensure_ascii=False, indent=2)};",
        "",
        f"export const OCCUPATION_GAP_DATA = {json.dumps(occupation_data, ensure_ascii=False, indent=2)};",
        "",
        f"export const INDUSTRY_GAP_DATA = {json.dumps(industry_data, ensure_ascii=False, indent=2)};",
        "",
        f"export const REGION_GAP_DATA = {json.dumps(region_data, ensure_ascii=False, indent=2)};",
        "",
        f"export const SECTOR_GAP_DATA = {json.dumps(sector_data, ensure_ascii=False, indent=2)};",
        "",
        f"export const AGE_OCCUPATION_GAP_DATA = {json.dumps(age_occupation_data, ensure_ascii=False, indent=2)};",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
