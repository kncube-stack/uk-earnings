from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def compute_gap(male_median, female_median):
    if male_median in (None, 0) or female_median is None:
        return None
    return round(((male_median - female_median) / male_median) * 100, 1)


def build_detail_rows(path: Path, male_sheet_name: str, female_sheet_name: str):
    male_wb = load_workbook(path, data_only=True)
    female_wb = male_wb
    male_sheet = male_wb[male_sheet_name]
    female_sheet = female_wb[female_sheet_name]
    grouped = {}

    for row in male_sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        row_id = str(row_id).strip() if row_id is not None else ""
        if len(row_id) != 4 or not row_id.isdigit():
            continue
        grouped.setdefault(row_id[0], {})[row_id] = {
            "label": " ".join(str(row[0]).split()),
            "maleMedian": clean_value(row[3]),
        }

    for row in female_sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        row_id = str(row_id).strip() if row_id is not None else ""
        if len(row_id) != 4 or not row_id.isdigit():
            continue
        entry = grouped.setdefault(row_id[0], {}).setdefault(row_id, {"label": " ".join(str(row[0]).split())})
        entry["femaleMedian"] = clean_value(row[3])

    output = {}
    for group_id, rows in grouped.items():
        output[group_id] = [
            {
                "id": row_id,
                "majorGroupId": group_id,
                "label": payload["label"],
                "shortLabel": row_id,
                "maleMedian": payload.get("maleMedian"),
                "femaleMedian": payload.get("femaleMedian"),
                "median": round((payload["maleMedian"] + payload["femaleMedian"]) / 2, 2)
                if payload.get("maleMedian") is not None and payload.get("femaleMedian") is not None
                else None,
                "gapPct": compute_gap(payload.get("maleMedian"), payload.get("femaleMedian")),
            }
            for row_id, payload in rows.items()
        ]
    return output


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_gender_gap_detail_data.py <table14_dir>")

    source_dir = Path(sys.argv[1])
    path = source_dir / "PROV - Occupation SOC20 (4) Table 14.6a   Hourly pay - Excluding overtime 2025.xlsx"

    output = {
        "all": build_detail_rows(path, "Male", "Female"),
        "ft": build_detail_rows(path, "Male Full-Time", "Female Full-Time"),
    }

    sys.stdout.write(
        "\n".join(
            [
                "// Generated from ONS ASHE 2025 Table 14 hourly excluding overtime workbook.",
                f"export const OCCUPATION_DETAIL_GAP_DATA = {json.dumps(output, ensure_ascii=False, indent=2)};",
                "",
            ]
        )
    )


if __name__ == "__main__":
    main()
