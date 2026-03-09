from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PERIOD_FILES = {
    "annual": "PROV - Work Region Occupation SOC20 (2) Table 3.7a   Annual pay - Gross 2025.xlsx",
    "weekly": "PROV - Work Region Occupation SOC20 (2) Table 3.1a   Weekly pay - Gross 2025.xlsx",
    "hourly": "PROV - Work Region Occupation SOC20 (2) Table 3.5a   Hourly pay - Gross 2025.xlsx",
    "hours": "PROV - Work Region Occupation SOC20 (2) Table 3.9a   Paid hours worked - Total 2025.xlsx",
}

SHEET_MAP = {
    "All": "all",
    "Male": "male",
    "Female": "female",
    "Full-Time": "ft",
    "Male Full-Time": "male_ft",
    "Female Full-Time": "female_ft",
}

OCCUPATION_SHORT_LABELS = {
    "1": "Managers",
    "2": "Professional",
    "3": "Associate prof.",
    "4": "Admin & secretarial",
    "5": "Skilled trades",
    "6": "Care & leisure",
    "7": "Sales & service",
    "8": "Drivers / operatives",
    "9": "Elementary",
}

OCCUPATION_LABEL_OVERRIDES = {
    "8": "Drivers, operatives and machine roles",
}

REGION_SHORT_LABELS = {
    "K02000001": "UK",
    "E12000001": "North East",
    "E12000002": "North West",
    "E12000003": "Yorks & Humber",
    "E12000004": "East Midlands",
    "E12000005": "West Midlands",
    "E12000006": "East of England",
    "E12000007": "London",
    "E12000008": "South East",
    "E12000009": "South West",
    "W92000004": "Wales",
    "S92000003": "Scotland",
    "N92000002": "N. Ireland",
}

FIELDS = [
    ("median", 4),
    ("mean", 6),
    ("p10", 8),
    ("p20", 9),
    ("p25", 10),
    ("p30", 11),
    ("p40", 12),
    ("p60", 13),
    ("p70", 14),
    ("p75", 15),
    ("p80", 16),
    ("p90", 17),
]


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_label(label, region_label):
    text = " ".join(str(label).split())
    if text == region_label:
        return region_label
    prefix = f"{region_label},"
    if text.startswith(prefix):
        return text[len(prefix):].strip()
    return text


def parse_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows_by_region = {}
    current_region_id = None
    current_region_label = None

    for row in sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        if row_id is None:
            continue
        row_id = str(row_id).strip()

        if row_id in REGION_SHORT_LABELS:
            current_region_id = row_id
            current_region_label = clean_value(row[0])
            rows_by_region.setdefault(current_region_id, [])
            continue

        if current_region_id is None or len(row_id) != 1 or not row_id.isdigit():
            continue

        label = normalize_label(row[0], current_region_label)
        rows_by_region[current_region_id].append(
            {
                "id": row_id,
                "label": OCCUPATION_LABEL_OVERRIDES.get(row_id, label),
                "shortLabel": OCCUPATION_SHORT_LABELS.get(row_id, label),
                **{field: clean_value(row[index - 1]) for field, index in FIELDS},
            }
        )

    return rows_by_region


def build_dataset(source_dir: Path):
    dataset = {}

    for period, filename in PERIOD_FILES.items():
        dataset[period] = {}
        workbook_path = source_dir / filename
        for sheet_name, key in SHEET_MAP.items():
            dataset[period][key] = parse_sheet(workbook_path, sheet_name)

    return dataset


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_region_occupation_data.py <extracted_table3_dir>")

    source_dir = Path(sys.argv[1])
    dataset = build_dataset(source_dir)

    output = [
        "// Generated from ONS ASHE 2025 Table 3 workbooks.",
        f"export const OCCUPATION_REGION_DATA = {json.dumps(dataset, ensure_ascii=False, indent=2)};",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
