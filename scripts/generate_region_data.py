from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PERIOD_FILES = {
    "annual": "PROV - Work Region Age Table WGOR Age.7a   Annual pay - Gross 2025.xlsx",
    "weekly": "PROV - Work Region Age Table WGOR Age.1a   Weekly pay - Gross 2025.xlsx",
    "hourly": "PROV - Work Region Age Table WGOR Age.5a   Hourly pay - Gross 2025.xlsx",
    "hours": "PROV - Work Region Age Table WGOR Age.9a   Paid hours worked - Total 2025.xlsx",
}

SHEET_MAP = {
    "All": "all",
    "Male": "male",
    "Female": "female",
    "Full-Time": "ft",
    "Male Full-Time": "male_ft",
    "Female Full-Time": "female_ft",
}

SHORT_LABELS = {
    "K02000001": "UK",
    "E12000001": "North East",
    "E12000002": "North West",
    "E12000003": "Yorks & Humber",
    "E12000004": "East Midlands",
    "E12000005": "West Midlands",
    "E12000006": "East of England",
    "E12000007": "London",
    "E12000008": "South East",
    "E12000009": "South West",
    "W92000004": "Wales",
    "S92000003": "Scotland",
    "N92000002": "N. Ireland",
}

FIELDS = [
    ("median", 4),
    ("mean", 6),
    ("p10", 8),
    ("p20", 9),
    ("p25", 10),
    ("p30", 11),
    ("p40", 12),
    ("p60", 13),
    ("p70", 14),
    ("p75", 15),
    ("p80", 16),
    ("p90", 17),
]


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows = []

    for row in sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        code = clean_value(row[1])
        code_str = str(code).strip() if code is not None else ""
        if not code_str or not code_str[0].isalpha():
            continue

        rows.append(
            {
                "id": code_str,
                "label": label,
                "shortLabel": SHORT_LABELS.get(code_str, label),
                **{field: clean_value(row[index - 1]) for field, index in FIELDS},
            }
        )

    return rows


def build_dataset(source_dir: Path):
    dataset = {}

    for period, filename in PERIOD_FILES.items():
        dataset[period] = {}
        workbook_path = source_dir / filename
        for sheet_name, key in SHEET_MAP.items():
            dataset[period][key] = parse_sheet(workbook_path, sheet_name)

    return dataset


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_region_data.py <extracted_table15_dir>")

    source_dir = Path(sys.argv[1])
    dataset = build_dataset(source_dir)

    output = [
        "// Generated from ONS ASHE 2025 Table 15 workbooks.",
        f"export const REGION_DATA = {json.dumps(dataset, ensure_ascii=False, indent=2)};",
        "",
        "export const REGION_OPTIONS = REGION_DATA.annual.all.map(({ id, label, shortLabel }) => ({",
        "  id,",
        "  label,",
        "  shortLabel,",
        "}));",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
