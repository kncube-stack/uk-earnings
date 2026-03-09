from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PERIOD_FILES = {
    "annual": "PROV - Occupation SOC20 (4) Table 14.7a   Annual pay - Gross 2025.xlsx",
    "weekly": "PROV - Occupation SOC20 (4) Table 14.1a   Weekly pay - Gross 2025.xlsx",
    "hourly": "PROV - Occupation SOC20 (4) Table 14.5a   Hourly pay - Gross 2025.xlsx",
    "hours": "PROV - Occupation SOC20 (4) Table 14.9a   Paid hours worked - Total 2025.xlsx",
}

SHEET_MAP = {
    "All": "all",
    "Male": "male",
    "Female": "female",
    "Full-Time": "ft",
    "Male Full-Time": "male_ft",
    "Female Full-Time": "female_ft",
}

FIELDS = [
    ("median", 4),
    ("mean", 6),
    ("p10", 8),
    ("p20", 9),
    ("p25", 10),
    ("p30", 11),
    ("p40", 12),
    ("p60", 13),
    ("p70", 14),
    ("p75", 15),
    ("p80", 16),
    ("p90", 17),
]


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def clean_label(label):
    return " ".join(str(label).split())


def parse_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    groups = {}

    for row in sheet.iter_rows(min_row=6, values_only=True):
        code = clean_value(row[1])
        code_str = str(code).strip() if code is not None else ""
        if len(code_str) != 4 or not code_str.isdigit():
            continue

        label = clean_label(row[0])
        major_group_id = code_str[0]
        groups.setdefault(major_group_id, []).append(
            {
                "id": code_str,
                "majorGroupId": major_group_id,
                "label": label,
                "shortLabel": code_str,
                **{field: clean_value(row[index - 1]) for field, index in FIELDS},
            }
        )

    return groups


def build_dataset(source_dir: Path):
    dataset = {}

    for period, filename in PERIOD_FILES.items():
        dataset[period] = {}
        workbook_path = source_dir / filename
        for sheet_name, key in SHEET_MAP.items():
            dataset[period][key] = parse_sheet(workbook_path, sheet_name)

    return dataset


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_occupation_detail_data.py <extracted_table14_dir>")

    source_dir = Path(sys.argv[1])
    dataset = build_dataset(source_dir)

    output = [
        "// Generated from ONS ASHE 2025 Table 14 workbooks.",
        f"export const OCCUPATION_DETAIL_DATA = {json.dumps(dataset, ensure_ascii=False, indent=2)};",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
