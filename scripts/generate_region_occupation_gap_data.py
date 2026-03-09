from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_FILE = "PROV - Work Region Occupation SOC20 (2) Table 3.6a   Hourly pay - Excluding overtime 2025.xlsx"

OCCUPATION_SHORT_LABELS = {
    "1": "Managers",
    "2": "Professional",
    "3": "Associate prof.",
    "4": "Admin & secretarial",
    "5": "Skilled trades",
    "6": "Care & leisure",
    "7": "Sales & service",
    "8": "Drivers / operatives",
    "9": "Elementary",
}

OCCUPATION_LABEL_OVERRIDES = {
    "8": "Drivers, operatives and machine roles",
}

REGION_SHORT_LABELS = {
    "K02000001": "UK",
    "E12000001": "North East",
    "E12000002": "North West",
    "E12000003": "Yorks & Humber",
    "E12000004": "East Midlands",
    "E12000005": "West Midlands",
    "E12000006": "East of England",
    "E12000007": "London",
    "E12000008": "South East",
    "E12000009": "South West",
    "W92000004": "Wales",
    "S92000003": "Scotland",
    "N92000002": "N. Ireland",
}


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_label(label, region_label):
    text = " ".join(str(label).split())
    if text == region_label:
        return region_label
    prefix = f"{region_label},"
    if text.startswith(prefix):
        return text[len(prefix):].strip()
    return text


def compute_gap(male_median, female_median):
    if male_median in (None, 0) or female_median is None:
        return None
    return round(((male_median - female_median) / male_median) * 100, 1)


def make_gap_row(row_id, label, male_median, female_median):
    overall_median = None
    if male_median is not None and female_median is not None:
        overall_median = round((male_median + female_median) / 2, 2)
    return {
        "id": row_id,
        "label": OCCUPATION_LABEL_OVERRIDES.get(row_id, label),
        "shortLabel": OCCUPATION_SHORT_LABELS.get(row_id, label),
        "maleMedian": male_median,
        "femaleMedian": female_median,
        "median": overall_median,
        "gapPct": compute_gap(male_median, female_median),
    }


def parse_sheet(sheet):
    rows_by_region = {}
    current_region_id = None
    current_region_label = None

    for row in sheet.iter_rows(min_row=6, values_only=True):
        row_id = clean_value(row[1])
        if row_id is None:
            continue
        row_id = str(row_id).strip()

        if row_id in REGION_SHORT_LABELS:
            current_region_id = row_id
            current_region_label = clean_value(row[0])
            rows_by_region.setdefault(current_region_id, {})
            continue

        if current_region_id is None or len(row_id) != 1 or not row_id.isdigit():
            continue

        rows_by_region[current_region_id][row_id] = {
            "label": normalize_label(row[0], current_region_label),
            "median": clean_value(row[3]),
        }

    return rows_by_region


def build_gap_dataset(source_dir: Path):
    workbook = load_workbook(source_dir / WORKBOOK_FILE, data_only=True)

    output = {}
    for key, male_sheet_name, female_sheet_name in [
        ("all", "Male", "Female"),
        ("ft", "Male Full-Time", "Female Full-Time"),
    ]:
        male_rows = parse_sheet(workbook[male_sheet_name])
        female_rows = parse_sheet(workbook[female_sheet_name])
        output[key] = {}
        for region_id, region_rows in male_rows.items():
            if region_id not in female_rows:
                continue
            output[key][region_id] = [
                make_gap_row(
                    row_id,
                    payload["label"],
                    payload["median"],
                    female_rows[region_id].get(row_id, {}).get("median"),
                )
                for row_id, payload in region_rows.items()
                if row_id in female_rows[region_id]
            ]

    return output


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_region_occupation_gap_data.py <extracted_table3_dir>")

    source_dir = Path(sys.argv[1])
    dataset = build_gap_dataset(source_dir)

    output = [
        "// Generated from ONS ASHE 2025 Table 3 hourly excluding overtime workbook.",
        f"export const OCCUPATION_REGION_GAP_DATA = {json.dumps(dataset, ensure_ascii=False, indent=2)};",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
