from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


PERIOD_FILES = {
    "annual": "PROV - PubPriv Table 13.7a   Annual pay - Gross 2025.xlsx",
    "weekly": "PROV - PubPriv Table 13.1a   Weekly pay - Gross 2025.xlsx",
    "hourly": "PROV - PubPriv Table 13.5a   Hourly pay - Gross 2025.xlsx",
    "hours": "PROV - PubPriv Table 13.9a   Paid hours worked - Total 2025.xlsx",
}

SHEET_MAP = {
    "All": "all",
    "Male": "male",
    "Female": "female",
    "Full-Time": "ft",
    "Male Full-Time": "male_ft",
    "Female Full-Time": "female_ft",
}

SHORT_LABELS = {
    "public-sector": "Public",
    "private-sector": "Private",
    "non-profit-body-or-mutual-association": "Non-profit",
    "not-classified": "Unclassified",
}

FIELDS = [
    ("median", 4),
    ("mean", 6),
    ("p10", 8),
    ("p20", 9),
    ("p25", 10),
    ("p30", 11),
    ("p40", 12),
    ("p60", 13),
    ("p70", 14),
    ("p75", 15),
    ("p80", 16),
    ("p90", 17),
]


def clean_value(value):
    if value in (None, "", "x", "..", ":", "-"):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "x", "..", ":", "-"}:
            return None
        try:
            numeric = float(stripped)
        except ValueError:
            return stripped
        if numeric.is_integer():
            return int(numeric)
        return round(numeric, 2)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def slugify(label: str):
    return re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")


def parse_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows = []

    for row in sheet.iter_rows(min_row=6, values_only=True):
        label = clean_value(row[0])
        if not isinstance(label, str):
            continue
        if label.startswith(("a  ", "b  ", "c  ", "KEY", "The quality", "Source:")):
            continue

        row_id = slugify(label)
        if row_id not in SHORT_LABELS:
            continue

        rows.append(
            {
                "id": row_id,
                "label": label,
                "shortLabel": SHORT_LABELS[row_id],
                **{field: clean_value(row[index - 1]) for field, index in FIELDS},
            }
        )

    return rows


def build_dataset(source_dir: Path):
    dataset = {}

    for period, filename in PERIOD_FILES.items():
        dataset[period] = {}
        workbook_path = source_dir / filename
        for sheet_name, key in SHEET_MAP.items():
            dataset[period][key] = parse_sheet(workbook_path, sheet_name)

    return dataset


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate_sector_data.py <extracted_table13_dir>")

    source_dir = Path(sys.argv[1])
    dataset = build_dataset(source_dir)

    output = [
        "// Generated from ONS ASHE 2025 Table 13 workbooks.",
        f"export const SECTOR_DATA = {json.dumps(dataset, ensure_ascii=False, indent=2)};",
        "",
        "export const SECTOR_OPTIONS = SECTOR_DATA.annual.all.map(({ id, label, shortLabel }) => ({",
        "  id,",
        "  label,",
        "  shortLabel,",
        "}));",
        "",
    ]

    sys.stdout.write("\n".join(output))


if __name__ == "__main__":
    main()
